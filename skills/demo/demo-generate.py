#!/usr/bin/env python3
"""
with-AI 商談デモ: 5つのビジネス文書を並列生成

相手の会社・担当者・提案内容に合わせてカスタマイズされた文書を生成する。

Usage:
  python3 demo-generate.py \
    --company "株式会社タイム" \
    --contact "田中太郎" \
    --title "営業部長" \
    --email "tanaka@time.co.jp" \
    --plan "AIKOMONプラン" \
    --amount 250000 \
    --output-dir ~/Desktop/demo-タイム
"""

import argparse
import json
import os
import sys
import threading
import time
from datetime import datetime, timedelta

# ============================================================
# PDF生成（reportlab）
# ============================================================
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas as pdf_canvas

# ============================================================
# Excel生成（openpyxl）
# ============================================================
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 日本語フォント
_FONT_PATH = "/Library/Fonts/Arial Unicode.ttf"
_FONT_ALT = "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"
_font_file = _FONT_PATH if os.path.exists(_FONT_PATH) else _FONT_ALT
pdfmetrics.registerFont(TTFont("ArialUnicode", _font_file))
FONT = "ArialUnicode"

# with-AI 会社情報
WITHAI = {
    "name": "with-AI株式会社",
    "representative": "代表取締役 勝又海斗",
    "address": "東京都渋谷区千駄ヶ谷5-16-10-1105",
    "phone": "090-1296-4814",
    "email": "info@with-ai.jp",
    "bank_name": "GMOあおぞらネット銀行（0310）",
    "branch_name": "法人営業部（101）",
    "account_type": "普通",
    "account_number": "2489145",
    "account_holder": "ウィズエーアイ（カ",
}

results = {}
lock = threading.Lock()
NOW = datetime.now()
YEAR = NOW.year
MONTH = NOW.month
DAY = NOW.day

# 月末日を計算
import calendar
LAST_DAY = calendar.monthrange(YEAR, MONTH)[1]


def report(name, status):
    with lock:
        results[name] = status
        print(f"  {'✅' if status == 'OK' else '❌'} {name}")


def yen(amount):
    return f"¥{amount:,}"


# ============================================================
# 1. 請求書PDF
# ============================================================
def generate_invoice(args, output_dir):
    try:
        path = os.path.join(output_dir, f"請求書_{YEAR}{MONTH:02d}.pdf")
        tax = int(args.amount * 0.10)
        total = args.amount + tax

        c = pdf_canvas.Canvas(path, pagesize=A4)
        w, h = A4

        # タイトル
        c.setFont(FONT, 24)
        c.drawCentredString(w / 2, h - 40 * mm, "請 求 書")

        # 番号・日付
        c.setFont(FONT, 9)
        c.drawRightString(w - 20 * mm, h - 55 * mm, f"請求書番号: INV-{YEAR}{MONTH:02d}-001")
        c.drawRightString(w - 20 * mm, h - 61 * mm, f"請求日: {YEAR}年{MONTH}月{DAY}日")

        # 請求先
        c.setFont(FONT, 14)
        c.drawString(20 * mm, h - 60 * mm, f"{args.company} 御中")
        c.setLineWidth(0.5)
        c.line(20 * mm, h - 62 * mm, 110 * mm, h - 62 * mm)

        y = h - 70 * mm
        c.setFont(FONT, 9)
        if args.contact:
            label = args.contact
            if args.title:
                label = f"{args.contact}（{args.title}）"
            c.drawString(20 * mm, y, f"{label} 様")

        # 発行元
        rx, ry = 125 * mm, h - 75 * mm
        c.setFont(FONT, 11)
        c.drawString(rx, ry, WITHAI["name"])
        for i, line in enumerate([
            WITHAI["address"],
            f"TEL: {WITHAI['phone']}",
            f"Email: {WITHAI['email']}",
            WITHAI["representative"],
        ]):
            c.setFont(FONT, 8)
            c.drawString(rx, ry - (i + 1) * 5 * mm, line)

        # 合計金額
        y = h - 100 * mm
        c.setFont(FONT, 10)
        c.drawString(20 * mm, y, "ご請求金額（税込）")
        c.setFillColor(colors.HexColor("#f0f0f0"))
        c.rect(20 * mm, y - 14 * mm, 90 * mm, 12 * mm, fill=True, stroke=True)
        c.setFillColor(colors.black)
        c.setFont(FONT, 18)
        c.drawCentredString(65 * mm, y - 12 * mm, yen(total))

        # 支払期限
        y -= 22 * mm
        c.setFont(FONT, 10)
        c.drawString(20 * mm, y, f"お支払期限: {YEAR}年{MONTH}月{LAST_DAY}日")

        # 明細テーブル
        y -= 15 * mm
        tx = 20 * mm
        tw = w - 40 * mm
        cw = [tw * 0.55, tw * 0.15, tw * 0.15, tw * 0.15]

        c.setFillColor(colors.HexColor("#333333"))
        c.rect(tx, y - 8 * mm, tw, 8 * mm, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont(FONT, 9)
        for i, header in enumerate(["品目", "数量", "単価", "金額"]):
            cx = tx + sum(cw[:i]) + cw[i] / 2
            c.drawCentredString(cx, y - 6 * mm, header)

        c.setFillColor(colors.black)
        ry = y - 18 * mm
        c.setFont(FONT, 9)
        c.drawString(tx + 3 * mm, ry, f"{args.plan}（{YEAR}年{MONTH}月分）")
        c.drawCentredString(tx + cw[0] + cw[1] / 2, ry, "1")
        c.drawRightString(tx + cw[0] + cw[1] + cw[2] - 3 * mm, ry, yen(args.amount))
        c.drawRightString(tx + tw - 3 * mm, ry, yen(args.amount))

        # 小計・税・合計
        sy = ry - 15 * mm
        sx = tx + cw[0] + cw[1]
        for label, val in [("小計", args.amount), ("消費税（10%）", tax)]:
            c.setFont(FONT, 9)
            c.drawString(sx, sy, label)
            c.drawRightString(tx + tw - 3 * mm, sy, yen(val))
            sy -= 7 * mm
        c.setFont(FONT, 11)
        c.drawString(sx, sy, "合計（税込）")
        c.drawRightString(tx + tw - 3 * mm, sy, yen(total))

        # 振込先
        by = sy - 20 * mm
        c.setFont(FONT, 10)
        c.drawString(20 * mm, by, "お振込先")
        c.setLineWidth(0.5)
        c.line(20 * mm, by - 3 * mm, 110 * mm, by - 3 * mm)
        by -= 10 * mm
        c.setFont(FONT, 9)
        for line in [
            f"銀行名: {WITHAI['bank_name']}",
            f"支店名: {WITHAI['branch_name']}",
            f"口座種別: {WITHAI['account_type']}",
            f"口座番号: {WITHAI['account_number']}",
            f"口座名義: {WITHAI['account_holder']}",
        ]:
            c.drawString(25 * mm, by, line)
            by -= 6 * mm
        by -= 3 * mm
        c.setFont(FONT, 8)
        c.drawString(25 * mm, by, "※ 振込手数料はお客様のご負担でお願いいたします。")

        c.save()
        report("請求書PDF", "OK")
    except Exception as e:
        report("請求書PDF", f"ERROR: {e}")


# ============================================================
# 2. 経費台帳Excel
# ============================================================
def generate_expense(args, output_dir):
    try:
        path = os.path.join(output_dir, "経費台帳.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "経費台帳"

        hdr_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(bottom=Side(style="thin", color="CCCCCC"))

        ws.merge_cells("A1:F1")
        ws["A1"] = f"経費台帳 — {YEAR}年{MONTH}月"
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["日付", "カテゴリ", "内容", "金額", "支払方法", "備考"]
        widths = [14, 14, 35, 14, 16, 24]
        for i, (h, w_) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=3, column=i, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width = w_

        # 相手の会社名を使ったリアルな経費データ
        short = args.company.replace("株式会社", "").replace("合同会社", "").strip()
        contact = args.contact or "担当者"

        expenses = [
            (f"{YEAR}/04/01", "交通費", f"{short} 初回訪問（新宿→渋谷）", 380, "IC", ""),
            (f"{YEAR}/04/01", "交際費", f"{contact}様とランチミーティング", 3200, "クレジットカード", "渋谷 イタリアン"),
            (f"{YEAR}/04/03", "消耗品", "プレゼン用 HDMI アダプタ", 2480, "クレジットカード", "Amazon"),
            (f"{YEAR}/04/05", "交通費", f"{short} 提案プレゼン（渋谷→{short}様オフィス）", 580, "IC", "往復"),
            (f"{YEAR}/04/05", "交際費", f"{contact}様とカフェ打合せ", 1800, "現金", "スターバックス"),
            (f"{YEAR}/04/07", "通信費", "Zoom Pro 月額", 2200, "クレジットカード", "月額固定"),
            (f"{YEAR}/04/08", "交通費", f"{short} 契約打合せ（往復）", 760, "IC", ""),
            (f"{YEAR}/04/08", "交際費", f"{contact}様と契約祝いディナー", 12800, "クレジットカード", "恵比寿 和食"),
            (f"{YEAR}/04/10", "書籍", f"{short}業界リサーチ用書籍", 2980, "クレジットカード", "Amazon"),
            (f"{YEAR}/04/15", "交通費", f"{short} オンボーディング訪問", 380, "IC", ""),
        ]

        for i, row in enumerate(expenses, 4):
            for j, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = border
                if j == 4:
                    cell.number_format = '¥#,##0'
                    cell.alignment = Alignment(horizontal="right")

        # 合計行
        total_row = 4 + len(expenses)
        ws.cell(row=total_row, column=3, value="合計").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=4)
        total_cell.value = sum(e[3] for e in expenses)
        total_cell.number_format = '¥#,##0'
        total_cell.font = Font(bold=True, size=12)
        total_cell.alignment = Alignment(horizontal="right")

        # カテゴリ別集計シート
        ws2 = wb.create_sheet("カテゴリ別集計")
        ws2.merge_cells("A1:C1")
        ws2["A1"] = "カテゴリ別 経費集計"
        ws2["A1"].font = Font(bold=True, size=14)

        categories = {}
        counts = {}
        for e in expenses:
            categories[e[1]] = categories.get(e[1], 0) + e[3]
            counts[e[1]] = counts.get(e[1], 0) + 1

        for i, h in enumerate(["カテゴリ", "件数", "金額"], 1):
            cell = ws2.cell(row=3, column=i, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            ws2.column_dimensions[get_column_letter(i)].width = 18

        for i, (cat, amt) in enumerate(sorted(categories.items()), 4):
            ws2.cell(row=i, column=1, value=cat)
            ws2.cell(row=i, column=2, value=counts[cat])
            ws2.cell(row=i, column=3, value=amt).number_format = '¥#,##0'

        wb.save(path)
        report("経費台帳Excel", "OK")
    except Exception as e:
        report("経費台帳Excel", f"ERROR: {e}")


# ============================================================
# 3. 売上レポートExcel（グラフ付き）
# ============================================================
def generate_sales_report(args, output_dir):
    try:
        path = os.path.join(output_dir, "売上レポート.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "月次売上"

        hdr_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        new_fill = PatternFill(start_color="d4e6f1", end_color="d4e6f1", fill_type="solid")

        ws.merge_cells("A1:E1")
        ws["A1"] = f"月次売上レポート — {YEAR}年{MONTH}月"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A2:E2")
        ws["A2"] = f"作成日: {YEAR}年{MONTH}月{DAY}日"
        ws["A2"].font = Font(size=9, color="888888")

        headers = ["クライアント", "プラン", "月額（税抜）", "消費税", "合計（税込）"]
        widths = [24, 24, 16, 14, 16]
        for i, (h, w_) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width = w_

        # 既存クライアント（架空） + 今回の提案先
        clients = [
            ("株式会社アルファ", "AI相談プラン", 49800),
            ("株式会社ベータ", "AIKOMONプラン", 250000),
            ("株式会社ガンマ", "プレミアムプラン", 500000),
            (args.company, args.plan, args.amount),  # ← NEW
        ]

        for i, (name, plan, amount) in enumerate(clients, 5):
            tax = int(amount * 0.10)
            total = amount + tax
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=plan)
            ws.cell(row=i, column=3, value=amount).number_format = '¥#,##0'
            ws.cell(row=i, column=4, value=tax).number_format = '¥#,##0'
            ws.cell(row=i, column=5, value=total).number_format = '¥#,##0'
            # 新規クライアント（提案先）をハイライト
            if name == args.company:
                for j in range(1, 6):
                    ws.cell(row=i, column=j).fill = new_fill
                ws.cell(row=i, column=1, value=f"{name} ★NEW")

        # 合計行
        tr = 5 + len(clients)
        grand_sub = sum(c[2] for c in clients)
        grand_tax = sum(int(c[2] * 0.10) for c in clients)
        grand_total = grand_sub + grand_tax
        ws.cell(row=tr, column=1, value="合計").font = Font(bold=True, size=12)
        ws.cell(row=tr, column=3, value=grand_sub).number_format = '¥#,##0'
        ws.cell(row=tr, column=3).font = Font(bold=True)
        ws.cell(row=tr, column=4, value=grand_tax).number_format = '¥#,##0'
        ws.cell(row=tr, column=4).font = Font(bold=True)
        ws.cell(row=tr, column=5, value=grand_total).number_format = '¥#,##0'
        ws.cell(row=tr, column=5).font = Font(bold=True, size=12)

        # MRR推移（提案先が加わるとこう変わる）
        ws2 = wb.create_sheet("売上推移")
        ws2.merge_cells("A1:C1")
        ws2["A1"] = "月次売上推移（MRR）"
        ws2["A1"].font = Font(bold=True, size=14)
        ws2.merge_cells("A2:C2")
        ws2["A2"] = f"※ {args.company} 加入後の見込みを含む"
        ws2["A2"].font = Font(size=9, color="1a5276", italic=True)

        existing_mrr = sum(int(c[2] * 1.10) for c in clients[:3])
        new_mrr = grand_total

        for i, h in enumerate(["月", "売上（税込）"], 1):
            cell = ws2.cell(row=4, column=i, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            ws2.column_dimensions[get_column_letter(i)].width = 18

        months_data = []
        for offset in range(5, 0, -1):
            m = MONTH - offset
            y = YEAR
            if m <= 0:
                m += 12
                y -= 1
            months_data.append((f"{y}年{m}月", existing_mrr))
        months_data.append((f"{YEAR}年{MONTH}月", new_mrr))

        for i, (month, val) in enumerate(months_data, 5):
            ws2.cell(row=i, column=1, value=month)
            ws2.cell(row=i, column=2, value=val).number_format = '¥#,##0'

        # グラフ
        chart = BarChart()
        chart.type = "col"
        chart.title = "月次MRR推移"
        chart.y_axis.title = "売上（税込）"
        chart.style = 10
        chart.width = 20
        chart.height = 12
        data = Reference(ws2, min_col=2, min_row=4, max_row=10)
        cats = Reference(ws2, min_col=1, min_row=5, max_row=10)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = "1a5276"
        ws2.add_chart(chart, "A12")

        wb.save(path)
        report("売上レポートExcel", "OK")
    except Exception as e:
        report("売上レポートExcel", f"ERROR: {e}")


# ============================================================
# 4. コンタクト情報カードPDF
# ============================================================
def generate_card(args, output_dir):
    try:
        contact = args.contact or "担当者"
        path = os.path.join(output_dir, f"{contact}_コンタクト.pdf")
        c = pdf_canvas.Canvas(path, pagesize=A4)
        w, h = A4

        c.setFont(FONT, 20)
        c.drawCentredString(w / 2, h - 35 * mm, "コンタクト情報")

        # カード枠
        cx, cy = 25 * mm, h - 140 * mm
        cw_, ch = w - 50 * mm, 90 * mm
        c.setStrokeColor(colors.HexColor("#333333"))
        c.setLineWidth(1.5)
        c.roundRect(cx, cy, cw_, ch, 5 * mm, stroke=True, fill=False)

        # 会社名
        c.setFont(FONT, 10)
        c.setFillColor(colors.HexColor("#666666"))
        c.drawString(cx + 10 * mm, cy + ch - 15 * mm, args.company)

        # 名前
        c.setFillColor(colors.black)
        c.setFont(FONT, 22)
        c.drawString(cx + 10 * mm, cy + ch - 32 * mm, contact)

        # 役職
        if args.title:
            c.setFont(FONT, 11)
            c.setFillColor(colors.HexColor("#666666"))
            c.drawString(cx + 10 * mm, cy + ch - 43 * mm, args.title)

        # 区切り線
        line_y = cy + ch - 52 * mm
        c.setStrokeColor(colors.HexColor("#dddddd"))
        c.setLineWidth(0.5)
        c.line(cx + 10 * mm, line_y, cx + cw_ - 10 * mm, line_y)

        # 連絡先
        c.setFillColor(colors.black)
        c.setFont(FONT, 9)
        details = []
        if args.email:
            details.append(f"Email:  {args.email}")
        else:
            details.append("Email:  （未登録）")
        details.append("TEL:    （未登録）")
        details.append("住所:   （未登録）")

        for i, line in enumerate(details):
            c.drawString(cx + 10 * mm, line_y - (i + 1) * 7 * mm, line)

        # 登録情報セクション
        y = cy - 20 * mm
        c.setFont(FONT, 12)
        c.setFillColor(colors.black)
        c.drawString(25 * mm, y, "登録情報")
        c.setLineWidth(0.5)
        c.setStrokeColor(colors.black)
        c.line(25 * mm, y - 3 * mm, w - 25 * mm, y - 3 * mm)

        info = [
            ("登録日", f"{YEAR}年{MONTH:02d}月{DAY:02d}日"),
            ("ステータス", "新規コンタクト → 商談中"),
            ("提案内容", f"{args.plan}（月額 {yen(args.amount)}）"),
            ("担当", "勝又海斗"),
            ("次回アクション", "契約書送付 → オンボーディング"),
        ]

        y -= 12 * mm
        c.setFont(FONT, 9)
        for label, val in info:
            c.setFillColor(colors.HexColor("#666666"))
            c.drawString(30 * mm, y, label)
            c.setFillColor(colors.black)
            c.drawString(78 * mm, y, val)
            y -= 8 * mm

        c.save()
        report("コンタクトPDF", "OK")
    except Exception as e:
        report("コンタクトPDF", f"ERROR: {e}")


# ============================================================
# 5. フォローメール下書き
# ============================================================
def generate_email(args, output_dir):
    try:
        contact = args.contact or "ご担当者"
        path = os.path.join(output_dir, "フォローメール.txt")

        to_line = args.email if args.email else "（メールアドレス未登録）"
        title_line = ""
        if args.title:
            title_line = f"\n{args.company} {args.title}"

        content = f"""To: {to_line}
Subject: 【with-AI】本日はありがとうございました — {args.plan}のご提案について

{contact} 様{title_line}

お世話になっております。with-AI株式会社の勝又です。
本日はお忙しい中、お時間をいただきありがとうございました。

ご提案させていただいた「{args.plan}」について、
改めて概要をお送りいたします。

■ ご提案内容
  サービス:   {args.plan}
  月額:       {yen(args.amount)}（税抜）
  消費税:     {yen(int(args.amount * 0.10))}
  合計:       {yen(int(args.amount * 1.10))}（税込）

■ 導入までの流れ
  1. ご契約の合意 → 契約書締結
  2. オンボーディングミーティング（1時間程度）
  3. 初回ヒアリング → AI活用方針の策定
  4. 運用開始

ご不明点やご質問がございましたら、お気軽にご連絡ください。
{args.company}様のAI活用を全力でサポートさせていただきます。

今後ともよろしくお願いいたします。

――――――――――――――――――――
勝又 海斗
with-AI株式会社 代表取締役
Email: info@with-ai.jp
TEL: 090-1296-4814
――――――――――――――――――――
"""
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        report("フォローメール", "OK")
    except Exception as e:
        report("フォローメール", f"ERROR: {e}")


# ============================================================
# メイン: 5つを並列実行
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="with-AI 商談デモ")
    parser.add_argument("--company", required=True, help="相手の会社名")
    parser.add_argument("--contact", default="", help="担当者名")
    parser.add_argument("--title", default="", help="役職")
    parser.add_argument("--email", default="", help="メールアドレス")
    parser.add_argument("--plan", required=True, help="提案内容・サービス名")
    parser.add_argument("--amount", type=int, required=True, help="金額（税抜）")
    parser.add_argument("--output-dir", required=True, help="出力先")
    args = parser.parse_args()

    output_dir = os.path.expanduser(args.output_dir)
    os.makedirs(output_dir, exist_ok=True)

    print(f"\n━━━ with-AI デモ: 並列生成開始 ━━━")
    print(f"  会社:     {args.company}")
    if args.contact:
        label = args.contact
        if args.title:
            label += f"（{args.title}）"
        print(f"  担当者:   {label}")
    print(f"  提案:     {args.plan} — {yen(args.amount)}/月")
    print(f"  出力先:   {output_dir}")
    print()

    start = time.time()

    threads = [
        threading.Thread(target=generate_invoice, args=(args, output_dir)),
        threading.Thread(target=generate_expense, args=(args, output_dir)),
        threading.Thread(target=generate_sales_report, args=(args, output_dir)),
        threading.Thread(target=generate_card, args=(args, output_dir)),
        threading.Thread(target=generate_email, args=(args, output_dir)),
    ]

    for t in threads:
        t.start()
    for t in threads:
        t.join()

    elapsed = time.time() - start
    print(f"\n━━━ 完了（{elapsed:.1f}秒）━━━")
    print(f"  出力先: {output_dir}")

    errors = [k for k, v in results.items() if v != "OK"]
    if errors:
        print(f"\n  エラー: {', '.join(errors)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
